__author__ = "Ali Faruk Yucel"
__version__ = "1.0"
__status__ = "Production"
__maintainer__ = "Ali Faruk Yucel"

import re
import pyperclip
import openpyxl
import sys
import PyPDF2
import os

"This program takes in either a pdf file or uses the clipboard to find and extract phone numbers and emails" \
    "to an excel workbook."

# ? -> zero or one
# * -> zero or more
# + -> one or more
# {x} -> exactly x times
# {x,y} -> at least x at most y

# Common Character Classes
# \d any numeric digit from 0 to 9.
# \w any letter, numeric digit, or the underscore character 'word'
# \s any space, tab, or newline character 'space'
# Capital version are opposites. \D Any character that is NOT a numeric digit from 0 to 9.

os.chdir("C:\\Users\\Your-Path")  # Complete the path before usage

# TODO: Create a regex for phone numbers
phoneRegex = re.compile(r'''
(
(\+(\d\d\d)?)?      # country code
\s*                 # zero or more spaces
\d\s*\d\d           # area code
\s*
\d\d\s*\d\d\s*\d\d  # subscriber number
)
''', re.VERBOSE)

# TODO: Create a regex for email addresses
emailRegex = re.compile(r'''
[a-zA-Z0-9_.+]+          # name part
@                        # @ symbol
[a-zA-Z0-9_.+]+          # domain name part
''', re.VERBOSE)

text = ""
# TODO: Get the text off the clipboard
if len(sys.argv) == 1:
    text = pyperclip.paste()

# TODO: If given a pdf take the text by parsing the pdf.
else:
    pdf = open(sys.argv[1], 'rb')
    reader = PyPDF2.PdfFileReader(pdf)
    for pageNum in range(reader.numPages):
        page = reader.getPage(pageNum)
        text += "Page Number: " + str(pageNum + 1) + '\n' + page.extractText() + '\n'
    pdf.close()

# TODO: Extract the email/phone from this text
extractedPhones = phoneRegex.findall(text)
extractedEmails = emailRegex.findall(text)

allPhoneNumbers = []
for phoneNum in extractedPhones:
    allPhoneNumbers.append(phoneNum[0])

allEmails = []
for email in extractedEmails:
    allEmails.append(email)

# TODO: Copy the extracted email/phone to the clipboard. (If no pdf is given)
if len(sys.argv) == 1:
    allPhoneNumbersText = '\n'.join(allPhoneNumbers)
    allEmailsText = '\n'.join(allEmails)
    allThings = 'PHONE NUMBERS:\n' + allPhoneNumbersText + '\nEMAILS:\n' + allEmailsText
    pyperclip.copy(allThings)

# TODO: If a pdf was given as an argument, export phone numbers and emails as an excel workbook.
else:
    workbook = openpyxl.Workbook()
    phones = workbook.create_sheet()
    phones.title = "Phones"

    emails = workbook.create_sheet()
    emails.title = "Emails"

    counter = 0
    for i in range(len(allPhoneNumbers)):
        phones.cell(row=i + 1, column=1).value = allPhoneNumbers[i]
        counter += 1

    counter = 0
    for i in range(len(allEmails)):
        emails.cell(row=i + 1, column=1).value = allEmails[i]
        counter += 1

    workbook.save("Phones and Emails.xlsx")
